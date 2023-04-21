# Read excel file
import openpyxl
from datetime import datetime, time

# Define the path to the Excel file
excel_file = r"C:\Users\loren\Desktop\Test\BD_ORDENES.xlsx"

# Load the Excel file and select the first sheet
wb = openpyxl.load_workbook(excel_file)

# Create a list of dictionaries, where each dictionary represents a row in the Excel sheet
rows = []
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    row_dict = {}
    for idx, cell in enumerate(row):
        if isinstance(cell, datetime):
            cell = cell.timestamp()
        elif isinstance(cell, time):
            cell = cell.strftime('%H:%M:%S')
        row_dict[sheet.cell(row=1, column=idx+1).value] = cell
    rows.append(row_dict)

# Write to G. sheets
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from google.oauth2 import service_account

# Define G. sheets variables & credentials
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
KEY = 'key.json'
SPREADSHEET_ID ='1QeJpkDbajigV-RBy7Lkb70eM03nm-l0kwc-QpQT6ChQ'

creds = None
creds = service_account.Credentials.from_service_account_file(KEY, scopes=SCOPES)

service = build('sheets', 'v4', credentials=creds)
sheet = service.spreadsheets()

# Convert the list of dictionaries to a list of lists
data = []
headers = list(rows[0].keys())
#data.append(headers)
for row in rows:
    data.append([row[header] for header in headers])

try:
    # Insert the data into a new sheet in the Google Sheets document
    result = sheet.values().append(spreadsheetId=SPREADSHEET_ID,
                                    range='Hoja1!A1',
                                    valueInputOption='USER_ENTERED',
                                    body={'values': data}).execute()
except Exception as e:
    print(f"Error al insertar los datos: {e}")

print("Datos insertados correctamente")
