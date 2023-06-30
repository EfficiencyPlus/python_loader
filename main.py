def load_File():
    import openpyxl
    from datetime import datetime, time
    import os

    # Define variables
    log_file_path = os.path.join(os.path.dirname(__file__), "Upload_Logs.txt")
    excel_dir = os.path.dirname(os.path.abspath(__file__))
    tabname = "Cheques"
    rows = []
    SPREADSHEET_ID ='1iuqRmwP0j9xeynqYMsHkjpdcYXIQWKr8AfAI_Np8v40'
    KEY = {
    "type": "service_account",
    "project_id": "silent-polygon-384022",
    "private_key_id": "c5c1ab76fbc26df7948bfdcaf4919cc5f9250507",
    "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvQIBADANBgkqhkiG9w0BAQEFAASCBKcwggSjAgEAAoIBAQCoBqBlU8dzFw6T\n6oadDJaovRrzDr16VBPyNRmL8rBDiTfoUQx2Bf+S33g26Y3FdL3B9k4Wfwwdt0ap\noQk0pXBMtF9TuMLaApQEVeXXI5uc3uZhKdIPBMHqIanUtw4m2WtWWUV27TzUjp5B\nrS9XWw1fhCNAta9D7nYn/9LgEHN72IMk12ts0V2nt2Ymce/EjJbh0e0BFrunp2la\nSq/LwmooRKfh5FKKNr8ypkRymHOl19qpklqwwS6eqYY23ZmXsVV295EQZtTlJ0ZG\nyyjiep9RfU5NtyDRT+Wyuhyfcg+fMC1KJ3B3VaARb5mj+ag2eEBq+inYqsGj8tgq\nJa4mKJAXAgMBAAECggEAI5dTvmLiqpRFuDtDTFzwRMnZvYZvyEuq9fEdejFr+MJG\nH2WbmBxpKHlBg5AxQZQa5c0AeW9r55d81k2QH6CziuviDV/jlj5ZkBmJ1+po+e+V\nKRusrRW65Zc/3BjyXHRNcjrypeBWtEeJaiv/DSfySzs4JGrXtzH5wTtFNcUGUUMR\nk8vyjxiXeqBGdYPG76Wp115uDxKhnVXG0C8g/cJxl06Hpv8Vdtgch6N1b+uQJht1\nYZ6aBovzK9TZGZvFq0JrlHo8fOhELxZJ+JpUptwI2IDJSgyDLiVWo2mAha1pv0P8\nkiHmGKpLOhMGa4qwHHLhloOoKGLJcP3GLd0bH27lUQKBgQDTE9c+vjdb+Jqv/Kgt\nvBGORcve2ihKDPYeCGn3WcEWCbiL+CFHIdtv1+j+txWw+9wN2051AQif5kN4lGOE\ngmRgUgsNk2ok3z4JC6LLd5m7gQ17OiXgJV78V0wusaemqe7e+sbyjSKiFzlDAGmD\nrUOHI4Qco45ZLmm0ILakXMk3dQKBgQDLyTMaOL3PBa3vHQXB0owxfi5N7kR6PD+L\n5sXPxoNQUzaGQOazhLob6IPdG92jGU01a8DAFQl7gjfcBQKiSVcCz55XnTVRlFd/\naxleGKe6T8xUQnw+clJELzR3z10s3T/oSpxGGf3PUIMeRbbNQGHLEUuDj9xiMXco\nUmZtLKjD2wKBgQCCCzJM1LKUtljPXlFbRAho6DWPykaSQOQNYs+udhzQ8BJEIg6W\nnZRtls9UxwXVMYbE/r2atpVStnIRQxMBG7it4z4uEgmOSAAXYJTe2IPdksf0iMG3\nFC8uNraX+ho5Q28I7+ZMn4mb5cUAs6tOVhxJkSce4eGcrkyvjowYPZY21QKBgGbR\navEt88aCDRH5yK/UNVu7WE2FjBNIVp+VnfSJKjPa6EWwdTXaH7R5Ch8DMj5aQ7RS\nW3wOoSwptVlTRdLvwfDvI+r8rKwudj8ZoEDzm5zLKkLRELLJJ2yzjHtuqoo6T0Y6\nbsKH1qFE9ALQFb1VYIH+heUpLn50irRTbs5im/vFAoGAVkKsnkgGiX4L/qIsJVnP\n+L8QNDJI0hboBWFwKBGR+gj7R2K5C2Km3AyFAqXgvxWdfH/1F7xiFOEerWvhIWM9\nxVJSom6qLH0Y/VqUwqUxYmBZuSlC0xZdX9mwUEtoptk+rlo+jmf66E7s7cPLD3bW\nsDMhcSFsprmkAFHstGcE1cM=\n-----END PRIVATE KEY-----\n",
    "client_email": "python-test@silent-polygon-384022.iam.gserviceaccount.com",
    "client_id": "106247402746694320637",
    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
    "token_uri": "https://oauth2.googleapis.com/token",
    "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
    "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/python-test%40silent-polygon-384022.iam.gserviceaccount.com"
    }
    print("""  
********************************************************************
  _____ _____ _____ ___ ____ ___ _____ _   _  ______   __        
 | ____|  ___|  ___|_ _/ ___|_ _| ____| \ | |/ ___\ \ / /    _   
 |  _| | |_  | |_   | | |    | ||  _| |  \| | |    \ V /   _| |_ 
 | |___|  _| |  _|  | | |___ | || |___| |\  | |___  | |   |_   _|
 |_____|_|   |_|   |___\____|___|_____|_| \_|\____| |_|     |_|  
 
********************************************************************
 """)
    print(f"Leyendo datos del archivo de Excel desde {excel_dir}")

    try:
        for filename in os.listdir(excel_dir):
            if filename.endswith('.xlsx'):
                excel_file = os.path.join(excel_dir, filename)
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                for row in sheet.iter_rows(values_only=True):
                    row_dict = {}
                    for idx, cell in enumerate(row):
                        if idx != 2:
                            if isinstance(cell, (int, float)):
                                cell = str(cell)
                            else:
                                cell = str(cell)
                        else:
                            if isinstance(cell, datetime):
                                cell = cell.strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                cell = str(cell).split()[0]
                        if isinstance(cell, time):
                            cell = cell.strftime('%H:%M:%S')
                        row_dict[sheet.cell(row=1, column=idx + 1).value] = cell
                    rows.append(row_dict)

        print("Autenticando en Google Sheets")
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        from google.oauth2 import service_account

        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

        creds = service_account.Credentials.from_service_account_info(KEY, scopes=SCOPES)

        service = build('sheets', 'v4', credentials=creds)
        sheets_api = service.spreadsheets()

        data = []
        headers = list(rows[0].keys())
        for row in rows:
            data.append([row[header] for header in headers])

        result = sheets_api.values().get(spreadsheetId=SPREADSHEET_ID, range=tabname).execute()
        existing_data = result.get('values', [])

        existing_unique_ids = set()
        for row in existing_data:
            existing_unique_ids.add(row[1])

        new_data = []
        for row in data:
            if row[1] not in existing_unique_ids:
                new_data.append(row)

        if new_data:
            try:
                append_result = sheets_api.values().append(
                    spreadsheetId=SPREADSHEET_ID,
                    range=f'{tabname}!A1',
                    valueInputOption='USER_ENTERED',
                    insertDataOption='INSERT_ROWS',
                    body={'values': new_data}
                ).execute()
                rowNum = append_result['updates']['updatedRows']
                print(f"{rowNum} filas insertadas")
            except Exception as e:
                print(f"Error al insertar los datos: {e}")
                with open(log_file_path, "a") as log_file:
                    log_file.write(f"Execution Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: ")
                    log_file.write(f"Error al insertar los datos: {e}.\n")
                    log_file.write("\n")
            else:
                print("Datos insertados correctamente!")
                with open(log_file_path, "a") as log_file:
                    log_file.write(f"Execution Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: ")
                    log_file.write(f"{rowNum} filas insertadas\n")
                    log_file.write("\n")
        else:
            print("No hay datos nuevos para insertar")
            with open(log_file_path, "a") as log_file:
                log_file.write(f"Execution Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: ")
                log_file.write("No hay datos nuevos para insertar\n")
                log_file.write("\n")
    except Exception as e:
        print(f"Error al cargar el archivo de Excel: {e}")
        with open(log_file_path, "a") as log_file:
            log_file.write(f"Execution Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: ")
            log_file.write(f"Error al cargar el archivo de Excel: {e}.\n")
            log_file.write("\n")

    import time
    print("Esta ventana se cerrará en 3 segundos")
    time.sleep(3)

# Call the load_File function directly
load_File()
