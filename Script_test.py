from Send_Mail import send_mail
# Read excel file
import openpyxl
from datetime import datetime, time
import os

# Define the path to the Excel files directory
excel_dir = r"C:\Users\loren\Desktop\python_loader"

# Create a list of dictionaries, where each dictionary represents a row in the Excel sheet
rows = []

# Iterate over each file in the directory
# for filename in os.listdir(excel_dir):
#     if filename.endswith('.xlsx'): # Check if the file is an Excel file
#         excel_file = os.path.join(excel_dir, filename)
        
#         # Load the Excel file and select the first sheet
#         wb = openpyxl.load_workbook(excel_file)

#         sheet = wb.active
#         for row in sheet.iter_rows(values_only=True):
#             row_dict = {}
#             for idx, cell in enumerate(row):
#                 if idx != 2:  # Exclude column C
#                     if isinstance(cell, (int, float)):
#                         cell = str(cell)
#                 else:
#                     if isinstance(cell, datetime):
#                         cell = cell.strftime('%m/%d/%Y')
#                     else:
#                         # Handle the case if the cell value is not a datetime object
#                         cell = str(cell).split()[0]
#                 if isinstance(cell, time):
#                     cell = cell.strftime('%H:%M:%S')
#                 row_dict[sheet.cell(row=1, column=idx + 1).value] = cell
#             rows.append(row_dict)
for filename in os.listdir(excel_dir):
    if filename.endswith('.xlsx'): # Check if the file is an Excel file
        excel_file = os.path.join(excel_dir, filename)
        
        # Load the Excel file and select the first sheet
        wb = openpyxl.load_workbook(excel_file)

        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            row_dict = {}
            for idx, cell in enumerate(row):
                if idx != 2:  # Exclude column C
                    if isinstance(cell, (int, float)):
                        cell = str(cell)
                    else:
                        cell = str(cell)
                else:
                    if isinstance(cell, datetime):
                        cell = cell.strftime('%m/%d/%Y')
                    else:
                        # Handle the case if the cell value is not a datetime object
                        cell = str(cell).split()[0]
                if isinstance(cell, time):
                    cell = cell.strftime('%H:%M:%S')
                row_dict[sheet.cell(row=1, column=idx + 1).value] = cell
            rows.append(row_dict)


# Write to G. sheets
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from google.oauth2 import service_account

# Define G. sheets variables & credentials
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
KEY = 'key.json'
SPREADSHEET_ID ='1iuqRmwP0j9xeynqYMsHkjpdcYXIQWKr8AfAI_Np8v40' #'1QeJpkDbajigV-RBy7Lkb70eM03nm-l0kwc-QpQT6ChQ'

creds = None
creds = service_account.Credentials.from_service_account_file(KEY, scopes=SCOPES)

service = build('sheets', 'v4', credentials=creds)
sheets_api = service.spreadsheets() # Rename the variable to avoid collision with sheet data

# Convert the list of dictionaries to a list of lists
data = []
headers = list(rows[0].keys())
for row in rows:
    data.append([row[header] for header in headers])

# Retrieve the existing data from the sheet
result = sheets_api.values().get(spreadsheetId=SPREADSHEET_ID, range='Cheques').execute()       
existing_data = result.get('values', [])

# Create a set of the existing unique IDs
existing_unique_ids = set()
for row in existing_data:
    existing_unique_ids.add(row[1])

# Insert new data into the sheet
new_data = []
for row in data:
    if row[1] not in existing_unique_ids:
        new_data.append(row)

if new_data:
    try:
        # Insert the new data into a new sheet in the Google Sheets document
        append_result = sheets_api.values().append(spreadsheetId=SPREADSHEET_ID,
                                range='Cheques!A1',
                                valueInputOption='USER_ENTERED',
                                insertDataOption='INSERT_ROWS',
                                body={'values': new_data}).execute()
        rowNum = append_result.get('updates').get('updatedRows')
        print(f"{rowNum} filas insertadas.")
        #send_mail(rowNum)

    except Exception as e:
        print(f"Error al insertar los datos: {e}")

    print("Datos insertados correctamente")
else:
    print("No hay datos nuevos para insertar")
